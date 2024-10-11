from openpyxl import load_workbook
import inquirer
import os

from datetime import date

today = date.today()

datum = today.strftime("%Y.%m.%d")

file = "template.tex"
with open(file, 'r') as r:
    data = r.read()

    class bcolors:
        HEADER = '\033[95m'
        OKBLUE = '\033[94m'
        OKCYAN = '\033[96m'
        OKGREEN = '\033[92m'
        WARNING = '\033[93m'
        FAIL = '\033[91m'
        ENDC = '\033[0m'
        BOLD = '\033[1m'
        UNDERLINE = '\033[4m'

    error = True
    while error == True:
        try:
            wt = load_workbook("/home/nic/Dokumenty/1.FILE/Sešit1.xlsx")
            wt.save("/home/nic/Dokumenty/1.FILE/Sešit1.xlsx")
            error = False
            print(f"{bcolors.OKGREEN}excel OK{bcolors.ENDC}")

        except PermissionError as e:
            print(type(e).__name__)
            input("entr pro znovu načtení")
            error = True
                  
            
    wb = load_workbook(filename = "/home/nic/Dokumenty/1.FILE/Sešit1.xlsx")
    ws = wb['List1']
    #print(ws['C1'].value)

    nabidka = []

    #range(start, stop, step)
    for i in range(1, 50, 5):
        #print(ws[f'C{i}'].value, ws[f'C{i+1}'].value)
        nabidka.append([i,ws[f'C{i}'].value,ws[f'C{i+1}'].value])
        
    #print(nabidka)

    questions = [inquirer.List("size",message="Firma?",choices=nabidka, ),]

    theme={    "Question": {
            "mark_color": "yellow",
            "brackets_color": "normal",
        },
        "List": {
            "selection_color": "cyan",
            "selection_cursor": "->"
        }
    }
    theme1 = inquirer.themes.load_theme_from_dict(theme)


    answers = inquirer.prompt(questions, theme=theme1)
    print(answers["size"])


    vyber = answers["size"]
    vyber = vyber[0]
    print(vyber)

    print(ws[f'C{vyber}'].value,f"\t",ws[f'F{vyber}'].value)
    print(ws[f'C{vyber+1}'].value,f"\t",ws[f'F{vyber+1}'].value)
    print(ws[f'C{vyber+2}'].value,f"\t",ws[f'F{vyber+2}'].value)
    print(ws[f'C{vyber+3}'].value,f"\t",ws[f'F{vyber+3}'].value)


    wb2 = load_workbook(filename = '/home/nic/WebAPP/data/data.xlsx')
    ws2 = wb2['nic']
    imput = str(ws[f'F{vyber}'].value)
    print(imput)
    #tvorba indexu 1ft6108a-8 1ft6108-8af71-1ca0
    index = imput[0:6]+imput[15]+imput[7:9]
    if index.startswith("1PH7") or index.startswith("1PH8"):
        index = imput[0:4]+imput[9]
        
    if index.startswith("1FK2") or index.startswith("1FT2"):
        index = imput[0:4]+imput[15]
       
    print(index)

    #vyhledání indexu v tabulce
    for r in ws2.rows:
        if r[0].value == index:
            odmer = r[1].value

    try:
        odmer
    except NameError:
        #print("Variable is not defined....!")
        odmer = "?"
        print(odmer)
    else:
        #print("Variable is defined.")
        print(odmer)




    odporK = input("odpor na kostru? (00,0M/G) ")
    odporM = input("mezi z odpor? (00,0) ")
    odpor = "1820V/" + odporK + r"$\Omega$  " + odporM + r"$\Omega$"

    soupis = ["odporS","odporRS","mechS","meziU","mechR","elB","mechB","loz","mechO","sigO","dq","sil","sig","zadni","predni","elC","odporC","vent","znec"]

    for polozka in soupis:
        stav = input(polozka +" " + "stav? o,/,n ")
        if stav == "n":
            txt = r'\cellcolor{gray!50} \textcolor{red} {\textit{\textbf{NOK!}}}'
            print(r"\{\{" + polozka + r"\}\}")
            data = data.replace(r"\{\{" + polozka + r"\}\}", str(txt))
        elif stav == "/":
            txt = ('/')
            data = data.replace(r"\{\{" + polozka + r"\}\}", str(txt))
        else:
            txt = ('OK')
            data = data.replace(r"\{\{" + polozka + r"\}\}", str(txt))
    


#     odporS = input("stav? o/n ")
#     if odporS == "n":
#         odporS = r'\cellcolor{gray!50} \textcolor{red} {\textit{\textbf{NOK!}}}'
#         data = data.replace(r"\{\{odporS\}\}", str(odporS))
#     else:
#         odporS = ('OK')
#         data = data.replace(r"\{\{odporS\}\}", str(odporS))
#         
#     odporRS = input("stav? o/n ")
#     if odporRS == "n":
#         odporRS = r'\cellcolor{gray!50} \textcolor{red} {\textit{\textbf{NOK!}}}'
#         data = data.replace(r"\{\{odporRS\}\}", str(odporRS))
#     else:
#         odporRS = ('OK')
#         data = data.replace(r"\{\{odporRS\}\}", str(odporRS))
# 
#     meziU = input("stav? o/n ")
#     if meziU == "n":
#         meziU = r'\cellcolor{gray!50} \textcolor{red} {\textit{\textbf{NOK!}}}'
#         data = data.replace(r"\{\{meziU\}\}", str(meziU))
#     else:
#         meziU = ('OK')
#         data = data.replace(r"\{\{meziU\}\}", str(meziU))
# 
# #     def oknok(var):
# #         if var == "n":
# #         var = r'\cellcolor{gray!50} \textcolor{red} {\textit{\textbf{NOK!}}}'
# #         data = data.replace(r"\{\{odporS\}\}", str(var))
# #     else:
# #         var = ('OK')
# #         data = data.replace(r"\{\{odporS\}\}", str(var))
# 
    napeti = (input("gen napeti? (00,0) ") + "V")

    moment = (input("moment brzdy? (00,0) ") + "Nm")
    vel = input("velikost konektoru 1/1,5/2/3 ")
    kon = input("typ konektoru sro/nas ")
    sigK = input("sig. kabel? ")
    typ = input("typ tep. čidla KTY/PT1000 ")
    pozn = input("poznámka? ")


    data = data.replace(r"\{\{firma\}\}", str(ws[f'C{vyber}'].value))
    data = data.replace(r"\{\{akz\}\}", str(ws[f'C{vyber+1}'].value))
    data = data.replace(r"\{\{dbz\}\}", str(ws[f'C{vyber+2}'].value))
    data = data.replace(r"\{\{web\}\}", str(ws[f'C{vyber+3}'].value))
    
    data = data.replace(r"\{\{mlfb\}\}", str(ws[f'F{vyber}'].value))
    data = data.replace(r"\{\{z\}\}", str(ws[f'F{vyber+1}'].value))
    data = data.replace(r"\{\{cislo\}\}", str(ws[f'F{vyber+2}'].value))
    data = data.replace(r"\{\{datum\}\}", str(datum))
    
    data = data.replace(r"\{\{odpor\}\}", str(odpor))
    
    data = data.replace(r"\{\{moment\}\}", str(moment))
    
    data = data.replace(r"\{\{napeti\}\}", str(napeti))
    data = data.replace(r"\{\{odmer\}\}", str(odmer))
    data = data.replace(r"\{\{vel\}\}", str(vel))
    data = data.replace(r"\{\{typkon\}\}", str(kon))
    data = data.replace(r"\{\{sigK\}\}", str(sigK))
    data = data.replace(r"\{\{typ\}\}", str(typ))
    data = data.replace(r"\{\{pozn\}\}", str(pozn))
    obr = ""
    for foto in os.listdir("/home/nic/pdfdiag/oscil"):
        print(os.listdir("/home/nic/pdfdiag/oscil"))
        if foto.endswith('.png'):
            obr = obr + r"\begin{figure}[h!]%" + "\n" + r"\centering%" + "\n" + r"\includegraphics[width=550px]{/home/nic/pdfdiag/oscil/" + f"{foto}" + r"}%" + "\n" + r"\end{figure}" + "\n"
            
    obr = obr + r"\end{document}" + "\n"
    data = data.replace(r"\end{document}", obr)
    
with open(f'e-{file}', 'w') as o:
    o.write(data)
    
os.system(f"pdflatex e-{file}")

# nazev = str(datum) + "_" + str(ws[f'F{vyber}'].value) + "_" + str(ws[f'F{vyber+2}'].value)
# 
# with open(f'{nazev}.tex', 'w') as o:
#     o.write(data)
#     
# os.system(f"pdflatex {nazev}.tex")